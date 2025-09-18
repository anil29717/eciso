from app import app, db, Admin, Category, Question, Industry, PreRegisteredUser, User, GameSession, UserJourney
from flask import render_template, request, jsonify, redirect, url_for, session, send_file, flash
from flask_login import login_user, login_required, logout_user, current_user
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename
from datetime import datetime
import os
import json
import random
import csv
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image as ExcelImage
import io
import base64
from PIL import Image
import re

# Function to parse questions from all_industries_questions.txt
def parse_questions_from_file():
    """Parse all questions from all_industries_questions.txt file"""
    questions = []
    file_path = os.path.join(os.path.dirname(__file__), 'all_industries_questions.txt')
    
    if not os.path.exists(file_path):
        return questions
    
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            content = file.read()
        
        # Split by industry sections
        industry_sections = re.split(r'INDUSTRY:\s*([^\n]+)', content)
        
        # Process each industry section (skip first empty element)
        for i in range(1, len(industry_sections), 2):
            if i + 1 < len(industry_sections):
                industry_name = industry_sections[i].strip()
                industry_content = industry_sections[i + 1].strip()
                
                # Split questions by number pattern
                question_blocks = re.split(r'\n(?=\d+\.)', industry_content)
                
                for block in question_blocks:
                    if not block.strip():
                        continue
                    
                    lines = block.strip().split('\n')
                    if len(lines) < 6:  # Need at least question + 4 options + answer
                        continue
                    
                    # Parse question
                    question_line = lines[0]
                    question_match = re.match(r'\d+\.\s*(.+)', question_line)
                    if not question_match:
                        continue
                    
                    question_text = question_match.group(1).strip()
                    
                    # Parse options
                    options = {}
                    option_lines = []
                    correct_answer = None
                    
                    for line in lines[1:]:
                        line = line.strip()
                        if line.startswith(('A.', 'B.', 'C.', 'D.')):
                            option_letter = line[0]
                            option_text = line[2:].strip()
                            options[option_letter] = option_text
                            option_lines.append(line)
                        elif line.startswith('Correct Answer:'):
                            correct_answer = line.replace('Correct Answer:', '').strip()
                    
                    # Validate we have all required parts
                    if len(options) == 4 and correct_answer and question_text:
                        question_data = {
                            'id': len(questions) + 1,  # Generate unique ID
                            'industry': industry_name,
                            'question': question_text,
                            'options': options,
                            'correct_answer': correct_answer
                        }
                        questions.append(question_data)
    
    except Exception as e:
        print(f"Error parsing questions file: {e}")
    
    return questions

# Game Routes
@app.route('/')
def welcome():
    return render_template('game/welcome.html')

@app.route('/user-info')
def user_info():
    return render_template('game/user_info.html')

@app.route('/industry-select')
def industry_select():
    # Top 8 most frequent industries from bulk upload analysis
    top_industries = ['BFSI', 'Manufacturing', 'New-Age', 'IT/ITES', 'Healthcare/Pharma', 'Automotive', 'Conglomerate', 'Aviation']
    
    # Get highlighted industries based on frequency analysis
    highlighted_industries = []
    for industry_name in top_industries:
        industry = Industry.query.filter_by(name=industry_name).first()
        if industry:
            highlighted_industries.append(industry)
        else:
            # Create industry if it doesn't exist
            new_industry = Industry(name=industry_name, is_highlighted=True)
            db.session.add(new_industry)
            highlighted_industries.append(new_industry)
    
    # Get remaining industries
    other_industries = Industry.query.filter(~Industry.name.in_(top_industries)).all()
    
    # Commit any new industries
    db.session.commit()
    
    return render_template('game/industry_select.html',
                         highlighted_industries=highlighted_industries,
                         other_industries=other_industries)

@app.route('/game/selfie_capture')
def selfie_capture():
    return render_template('game/selfie_capture.html')

@app.route('/selfie-capture')
def selfie_capture_alt():
    return render_template('game/selfie_capture.html')

@app.route('/question')
def question():
    return render_template('game/question.html')

@app.route('/feedback')
def feedback():
    return render_template('game/feedback.html')

@app.route('/answer-feedback')
def answer_feedback():
    return render_template('game/feedback.html')

@app.route('/company-info')
def company_info():
    # Get industry from session, default to 'Technology' if not found
    industry = session.get('industry', 'Technology')
    return render_template('game/company_info.html', industry=industry)

@app.route('/industry-message')
def industry_message():
    return render_template('game/company_info.html')

@app.route('/thank-you')
def thank_you():
    return render_template('game/thank_you.html')

# API Routes
@app.route('/api/get-suggestions', methods=['POST'])
def get_suggestions():
    data = request.get_json()
    query = data.get('query', '').strip().lower()
    suggestion_type = data.get('type', 'name')  # 'name' or 'company'
    
    if not query or len(query) < 2:
        return jsonify({'suggestions': []})
    
    suggestions = []
    
    try:
        if suggestion_type == 'name':
            # Read names from bulk upload file
            names_file = os.path.join('bulk_uploads', 'names_20250914_230351.txt')
            if os.path.exists(names_file):
                with open(names_file, 'r', encoding='utf-8') as f:
                    names = [line.strip() for line in f.readlines() if line.strip()]
                
                # Filter names that match the query
                matching_names = [name for name in names if query in name.lower()]
                suggestions = matching_names[:10]  # Limit to 10 suggestions
        
        elif suggestion_type == 'company':
            # Read companies from bulk upload file
            companies_file = os.path.join('bulk_uploads', 'companies_20250914_230351.txt')
            if os.path.exists(companies_file):
                with open(companies_file, 'r', encoding='utf-8') as f:
                    companies = [line.strip() for line in f.readlines() if line.strip()]
                
                # Filter companies that match the query
                matching_companies = [company for company in companies if query in company.lower()]
                suggestions = matching_companies[:10]  # Limit to 10 suggestions
    
    except Exception as e:
        print(f"Error reading bulk upload files: {e}")
        return jsonify({'suggestions': []})
    
    return jsonify({'suggestions': suggestions})

@app.route('/api/check-user', methods=['POST'])
def check_user():
    data = request.get_json()
    name = data.get('name', '').strip().lower()
    
    if not name:
        return jsonify({'found': False})
    
    # Check if user exists in pre-registered users
    user = PreRegisteredUser.query.filter(
        PreRegisteredUser.name.ilike(f'%{name}%')
    ).first()
    
    if user:
        return jsonify({
            'found': True,
            'name': user.name,
            'company_name': user.company_name,
            'industry': user.industry
        })
    
    return jsonify({'found': False})

@app.route('/api/start-game', methods=['POST'])
def start_game():
    data = request.get_json()
    
    # Get user information from request
    name = data.get('name')
    company_name = data.get('company_name')
    industry = data.get('industry')
    email = data.get('email', '')
    phone = data.get('phone', '')
    job_title = data.get('job_title', '')
    department = data.get('department', '')
    
    # Check if this is an industry update for existing session
    if industry and not name and not company_name and 'journey_id' in session:
        try:
            # Update existing journey with industry
            journey = UserJourney.query.get(session['journey_id'])
            if journey:
                journey.industry = industry
                session['industry'] = industry
                db.session.commit()
                return jsonify({'success': True, 'message': 'Industry updated'})
        except Exception as e:
            print(f"Error updating industry: {e}")
            return jsonify({'success': False, 'message': 'Failed to update industry'}), 500
    
    # Validate required fields for new user creation
    if not name or not company_name:
        return jsonify({'success': False, 'message': 'Name and company name are required'}), 400
    
    # Store user info in session
    session['user_name'] = name
    session['company_name'] = company_name
    session['industry'] = industry
    session['game_start_time'] = datetime.utcnow().isoformat()
    
    # Save user information to database
    try:
        # Generate unique session ID
        import uuid
        session_id = str(uuid.uuid4())
        
        # Get client information
        ip_address = request.environ.get('HTTP_X_FORWARDED_FOR', request.environ.get('REMOTE_ADDR', 'unknown'))
        user_agent = request.headers.get('User-Agent', '')
        
        # Create new user record
        new_user = User(
            name=name,
            company_name=company_name,
            industry=industry,
            email=email,
            phone=phone,
            job_title=job_title,
            department=department,
            session_id=session_id,
            ip_address=ip_address,
            user_agent=user_agent
        )
        
        db.session.add(new_user)
        db.session.flush()  # Get the user ID without committing
        
        # Create new user journey record to track complete session
        journey_session_id = str(uuid.uuid4())
        new_journey = UserJourney(
            user_id=new_user.id,
            journey_session_id=journey_session_id,
            name=name,
            company_name=company_name,
            industry=industry,
            email=email,
            phone=phone,
            job_title=job_title,
            department=department,
            journey_start=datetime.utcnow(),
            is_completed=False
        )
        
        db.session.add(new_journey)
        db.session.commit()
        
        # Store user ID and journey ID in session for later reference
        session['user_id'] = new_user.id
        session['journey_id'] = new_journey.id
        session['journey_session_id'] = journey_session_id
        
        return jsonify({
            'success': True,
            'user_id': new_user.id,
            'message': 'User information saved successfully'
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Error saving user information: {e}")
        
        # Still allow game to continue even if database save fails
        return jsonify({
            'success': True,
            'message': 'Game started (user info stored in session)'
        })

@app.route('/api/get-question', methods=['GET'])
def get_question():
    # Get all questions from the text file
    all_questions = parse_questions_from_file()
    
    if not all_questions:
        return jsonify({'error': 'No questions available'}), 404
    
    # Get questions that haven't been asked to this user yet
    asked_questions = session.get('asked_questions', [])
    
    # Filter out already asked questions
    available_questions = [q for q in all_questions if q['id'] not in asked_questions]
    
    # If all questions have been asked, reset the list
    if not available_questions:
        available_questions = all_questions
        session['asked_questions'] = []
        asked_questions = []
    
    # Select a random question from available questions
    question = random.choice(available_questions)
    
    # Add this question to the asked questions list
    asked_questions.append(question['id'])
    session['asked_questions'] = asked_questions
    
    # Store question ID in session for answer validation
    session['current_question_id'] = question['id']
    session['current_question_correct_answer'] = question['correct_answer']
    
    # Convert correct answer letter to index for frontend
    answer_to_index = {'A': 0, 'B': 1, 'C': 2, 'D': 3}
    correct_answer_index = answer_to_index.get(question['correct_answer'], 0)
    
    response = jsonify({
        'id': question['id'],
        'question': question['question'],
        'options': question['options'],
        'correct_answer': correct_answer_index,
        'category': question['industry']
    })
    
    # Add cache-busting headers
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    
    return response

@app.route('/api/submit-answer', methods=['POST'])
def submit_answer():
    data = request.get_json()
    selected_answer_index = data.get('selected_answer')
    question_id = data.get('question_id')
    time_taken = data.get('time_taken', 30)
    is_timeout = data.get('is_timeout', False)
    
    # Convert index to letter (0->A, 1->B, 2->C, 3->D)
    answer_map = {0: 'A', 1: 'B', 2: 'C', 3: 'D'}
    selected_answer = answer_map.get(selected_answer_index, 'A')
    
    # Get correct answer from session (set in get_question)
    correct_answer = session.get('current_question_correct_answer')
    
    if not correct_answer:
        return jsonify({'error': 'Question session expired'}), 404
    
    is_correct = selected_answer == correct_answer
    
    # Get question text from file for display
    all_questions = parse_questions_from_file()
    question_text = "Question not found"
    for q in all_questions:
        if q['id'] == question_id:
            question_text = q['question']
            break
    
    # Store answer in session for feedback
    session['selected_answer'] = selected_answer
    session['correct_answer'] = correct_answer
    session['is_correct'] = is_correct
    session['question_text'] = question_text
    session['time_taken'] = time_taken
    
    # Save to database (using question_id as text since it's from file, not DB)
    try:
        game_session = GameSession(
            user_id=session.get('user_id'),
            journey_id=session.get('journey_id'),  # Link to complete user journey
            name=session.get('user_name', 'Anonymous'),
            company_name=session.get('company_name', 'Unknown'),
            industry=session.get('industry', 'Unknown'),
            question_id=question_id,  # This will be the file-based question ID
            selected_answer=selected_answer,
            is_correct=is_correct,
            selfie_filename=session.get('selfie_filename'),
            session_end=datetime.utcnow()
        )
        db.session.add(game_session)
        db.session.commit()
        print(f"Game session saved with selfie: {session.get('selfie_filename')}")
    except Exception as e:
        print(f"Error saving game session: {e}")
        db.session.rollback()
    
    return jsonify({
        'correct': is_correct,
        'correct_answer': question.correct_answer,
        'explanation': f"The correct answer is {question.correct_answer}"
    })

@app.route('/api/save-selfie', methods=['POST'])
def save_selfie():
    try:
        print("=== SAVE SELFIE REQUEST RECEIVED ===")
        data = request.get_json()
        print(f"Request data type: {type(data)}")
        
        if not data:
            print("ERROR: No JSON data received")
            return jsonify({'error': 'No JSON data received'}), 400
            
        image_data = data.get('image')
        print(f"Image data length: {len(image_data) if image_data else 0}")
        
        if not image_data:
            print("ERROR: No image data provided")
            return jsonify({'error': 'No image data provided'}), 400
        
        # Remove data URL prefix (data:image/jpeg;base64,)
        if ',' in image_data:
            image_data = image_data.split(',')[1]
            print("Removed data URL prefix")
        
        # Decode base64 image
        try:
            image_binary = base64.b64decode(image_data)
            print(f"Decoded image size: {len(image_binary)} bytes")
        except Exception as decode_error:
            print(f"ERROR: Failed to decode base64 image: {decode_error}")
            return jsonify({'error': 'Invalid image data format'}), 400
        
        # Ensure selfies directory exists
        selfies_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'selfies')
        print(f"Selfies directory: {selfies_dir}")
        if not os.path.exists(selfies_dir):
            os.makedirs(selfies_dir, exist_ok=True)
            print("Created selfies directory")
        
        # Generate filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        user_name = session.get('user_name') or 'user'
        user_name = user_name.replace(' ', '_').replace('/', '_')
        filename = f"selfie_{timestamp}_{user_name}.jpg"
        filepath = os.path.join(selfies_dir, filename)
        print(f"Saving to: {filepath}")
        
        # Save image
        try:
            with open(filepath, 'wb') as f:
                f.write(image_binary)
            print(f"File written successfully: {os.path.getsize(filepath)} bytes")
        except Exception as file_error:
            print(f"ERROR: Failed to write file: {file_error}")
            return jsonify({'error': 'Failed to write image file'}), 500
        
        # Store filename in session
        session['selfie_filename'] = filename
        print(f"Session updated with filename: {filename}")
        
        # Update UserJourney record with selfie filename
        try:
            journey_id = session.get('journey_id')
            if journey_id:
                journey = UserJourney.query.get(journey_id)
                if journey:
                    journey.selfie_filename = filename
                    db.session.commit()
                    print(f"UserJourney updated with selfie: {filename}")
        except Exception as journey_error:
            print(f"Warning: Failed to update UserJourney with selfie: {journey_error}")
            # Don't fail the request if journey update fails
        
        print(f"=== SELFIE SAVED SUCCESSFULLY: {filename} ===")
        return jsonify({'success': True, 'filename': filename})
    
    except Exception as e:
        print(f"=== ERROR SAVING SELFIE: {str(e)} ===")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': f'Failed to save selfie: {str(e)}'}), 500

@app.route('/api/complete-game', methods=['POST'])
def complete_game():
    try:
        # Mark UserJourney as completed when user reaches gift collection
        journey_id = session.get('journey_id')
        if journey_id:
            journey = UserJourney.query.get(journey_id)
            if journey:
                journey.is_completed = True
                journey.journey_end = datetime.utcnow()
                print(f"UserJourney {journey_id} marked as completed")
                # Commit the journey completion immediately to ensure it's saved
                db.session.commit()
        
        # Create final game session record if needed
        if session.get('current_question_id'):
            game_session = GameSession(
                user_id=session.get('user_id'),
                journey_id=journey_id,
                name=session.get('user_name'),
                company_name=session.get('company_name'),
                industry=session.get('industry'),
                question_id=session.get('current_question_id'),
                selected_answer=session.get('selected_answer'),
                is_correct=session.get('is_correct'),
                selfie_filename=session.get('selfie_filename'),
                session_end=datetime.utcnow()
            )
            
            db.session.add(game_session)
            db.session.commit()
        
        # Clear session
        session.clear()
        
        return jsonify({'success': True})
    
    except Exception as e:
        db.session.rollback()
        print(f"Error completing game: {e}")
        return jsonify({'error': str(e)}), 500

# Admin Routes
@app.route('/admin')
@app.route('/admin/dashboard')
def admin_dashboard():
    if not current_user.is_authenticated:
        return redirect(url_for('admin_login'))
    
    # Get statistics
    total_sessions = GameSession.query.count()
    total_questions = Question.query.count()
    total_categories = Category.query.count()
    total_users = PreRegisteredUser.query.count()
    
    return render_template('admin/dashboard.html',
                         total_sessions=total_sessions,
                         total_questions=total_questions,
                         total_categories=total_categories,
                         total_users=total_users)

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        # Handle JSON requests from frontend
        if request.is_json:
            data = request.get_json()
            username = data.get('username')
            password = data.get('password')
        else:
            # Handle form data requests
            username = request.form.get('username')
            password = request.form.get('password')
        
        if not username or not password:
            if request.is_json:
                return jsonify({'success': False, 'message': 'Username and password are required'}), 400
            else:
                flash('Username and password are required')
                return render_template('admin/login.html')
        
        admin = Admin.query.filter_by(username=username).first()
        
        if admin and check_password_hash(admin.password_hash, password):
            login_user(admin)
            if request.is_json:
                return jsonify({'success': True, 'message': 'Login successful'})
            else:
                return redirect(url_for('admin_dashboard'))
        else:
            if request.is_json:
                return jsonify({'success': False, 'message': 'Invalid username or password'}), 401
            else:
                flash('Invalid username or password')
    
    return render_template('admin/login.html')

@app.route('/admin/logout')
@login_required
def admin_logout():
    logout_user()
    return redirect(url_for('admin_login'))

@app.route('/admin/questions')
@login_required
def admin_questions():
    questions = Question.query.join(Category).all()
    categories = Category.query.all()
    return render_template('admin/questions.html', questions=questions, categories=categories)

@app.route('/admin/analytics')
@login_required
def admin_analytics():
    return render_template('admin/analytics.html')

@app.route('/admin/users')
@login_required
def admin_users():
    # Get both pre-registered users and actual game users
    pre_registered_users = PreRegisteredUser.query.all()
    game_users = User.query.order_by(User.created_at.desc()).all()
    industries = Industry.query.all()
    return render_template('admin/users.html', 
                         pre_registered_users=pre_registered_users,
                         game_users=game_users,
                         industries=industries)

@app.route('/admin/export-data')
@login_required
def export_data():
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Game Sessions Data"
    
    # Headers
    headers = ['Session ID', 'Name', 'Company', 'Industry', 'Email', 'Phone', 'Job Title', 
               'Department', 'Question', 'Selected Answer', 'Correct Answer', 'Is Correct', 
               'Selfie Filename', 'Session Date', 'IP Address', 'User Agent']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    sessions = GameSession.query.join(User, GameSession.user_id == User.id, isouter=True).all()
    for row, session in enumerate(sessions, 2):
        user = session.user if hasattr(session, 'user') and session.user else None
        
        ws.cell(row=row, column=1, value=session.id)
        ws.cell(row=row, column=2, value=session.name)
        ws.cell(row=row, column=3, value=session.company_name)
        ws.cell(row=row, column=4, value=session.industry)
        ws.cell(row=row, column=5, value=user.email if user else '')
        ws.cell(row=row, column=6, value=user.phone if user else '')
        ws.cell(row=row, column=7, value=user.job_title if user else '')
        ws.cell(row=row, column=8, value=user.department if user else '')
        ws.cell(row=row, column=9, value=session.question.question_text if session.question else '')
        ws.cell(row=row, column=10, value=session.selected_answer)
        ws.cell(row=row, column=11, value=session.question.correct_answer if session.question else '')
        ws.cell(row=row, column=12, value='Yes' if session.is_correct else 'No')
        ws.cell(row=row, column=13, value=session.selfie_filename or '')
        ws.cell(row=row, column=14, value=session.created_at.strftime('%Y-%m-%d %H:%M:%S'))
        ws.cell(row=row, column=15, value=user.ip_address if user else '')
        ws.cell(row=row, column=16, value=user.user_agent if user else '')
    
    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        as_attachment=True,
        download_name=f'game_sessions_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# API Routes for Admin

# Questions CRUD API
@app.route('/admin/api/questions', methods=['GET'])
@login_required
def get_questions():
    try:
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 10))
        search = request.args.get('search', '')
        category_id = request.args.get('category', '')
        
        query = Question.query
        
        # Apply search filter
        if search:
            query = query.filter(Question.question_text.contains(search))
        
        # Apply category filter
        if category_id:
            query = query.filter(Question.category_id == int(category_id))
        
        # Join with category for sorting and display
        query = query.join(Category).order_by(Question.created_at.desc())
        
        # Paginate results
        pagination = query.paginate(
            page=page, per_page=per_page, error_out=False
        )
        
        questions = [{
            'id': q.id,
            'category_id': q.category_id,
            'category_name': q.category.name,
            'question_text': q.question_text,
            'option_a': q.option_a,
            'option_b': q.option_b,
            'option_c': q.option_c,
            'option_d': q.option_d,
            'correct_answer': q.correct_answer,
            'created_at': q.created_at.isoformat()
        } for q in pagination.items]
        
        return jsonify({
            'success': True,
            'questions': questions,
            'pagination': {
                'page': page,
                'pages': pagination.pages,
                'per_page': per_page,
                'total': pagination.total
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/admin/api/questions/<int:question_id>', methods=['GET'])
@login_required
def get_question_by_id(question_id):
    try:
        question = Question.query.get_or_404(question_id)
        return jsonify({
            'success': True,
            'question': {
                'id': question.id,
                'category_id': question.category_id,
                'category_name': question.category.name,
                'question_text': question.question_text,
                'option_a': question.option_a,
                'option_b': question.option_b,
                'option_c': question.option_c,
                'option_d': question.option_d,
                'correct_answer': question.correct_answer,
                'created_at': question.created_at.isoformat()
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 404

@app.route('/admin/api/questions', methods=['POST'])
@login_required
def add_question():
    try:
        data = request.get_json()
        
        # Validate required fields
        required_fields = ['category_id', 'question_text', 'option_a', 'option_b', 'option_c', 'option_d', 'correct_answer']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'success': False, 'message': f'Missing required field: {field}'}), 400
        
        # Validate correct answer
        if data['correct_answer'] not in ['A', 'B', 'C', 'D']:
            return jsonify({'success': False, 'message': 'Correct answer must be A, B, C, or D'}), 400
        
        # Validate category exists
        category = Category.query.get(data['category_id'])
        if not category:
            return jsonify({'success': False, 'message': 'Invalid category'}), 400
        
        question = Question(
            category_id=data['category_id'],
            question_text=data['question_text'].strip(),
            option_a=data['option_a'].strip(),
            option_b=data['option_b'].strip(),
            option_c=data['option_c'].strip(),
            option_d=data['option_d'].strip(),
            correct_answer=data['correct_answer']
        )
        
        db.session.add(question)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Question added successfully',
            'question_id': question.id
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/admin/api/questions/<int:question_id>', methods=['PUT'])
@login_required
def update_question(question_id):
    try:
        question = Question.query.get_or_404(question_id)
        data = request.get_json()
        
        # Validate required fields
        required_fields = ['category_id', 'question_text', 'option_a', 'option_b', 'option_c', 'option_d', 'correct_answer']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'success': False, 'message': f'Missing required field: {field}'}), 400
        
        # Validate correct answer
        if data['correct_answer'] not in ['A', 'B', 'C', 'D']:
            return jsonify({'success': False, 'message': 'Correct answer must be A, B, C, or D'}), 400
        
        # Validate category exists
        category = Category.query.get(data['category_id'])
        if not category:
            return jsonify({'success': False, 'message': 'Invalid category'}), 400
        
        # Update question fields
        question.category_id = data['category_id']
        question.question_text = data['question_text'].strip()
        question.option_a = data['option_a'].strip()
        question.option_b = data['option_b'].strip()
        question.option_c = data['option_c'].strip()
        question.option_d = data['option_d'].strip()
        question.correct_answer = data['correct_answer']
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Question updated successfully'
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/admin/api/questions/<int:question_id>', methods=['DELETE'])
@login_required
def delete_question(question_id):
    try:
        question = Question.query.get_or_404(question_id)
        
        # Check if question is referenced in game sessions
        game_sessions = GameSession.query.filter_by(question_id=question_id).count()
        if game_sessions > 0:
            return jsonify({
                'success': False, 
                'message': f'Cannot delete question. It has been used in {game_sessions} game sessions.'
            }), 400
        
        db.session.delete(question)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Question deleted successfully'
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/admin/api/questions/bulk-import', methods=['POST'])
@login_required
def bulk_import_questions():
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'No file uploaded'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': 'No file selected'}), 400
        
        if not file.filename.lower().endswith(('.csv', '.xlsx', '.xls')):
            return jsonify({'success': False, 'message': 'Invalid file format. Please upload CSV or Excel file'}), 400
        
        # Save uploaded file temporarily
        filename = secure_filename(file.filename)
        temp_path = os.path.join('temp', filename)
        os.makedirs('temp', exist_ok=True)
        file.save(temp_path)
        
        imported_count = 0
        errors = []
        
        try:
            if filename.lower().endswith('.csv'):
                # Process CSV file
                with open(temp_path, 'r', encoding='utf-8') as csvfile:
                    reader = csv.DictReader(csvfile)
                    for row_num, row in enumerate(reader, start=2):
                        try:
                            imported_count += process_question_row(row, row_num)
                        except Exception as e:
                            errors.append(f'Row {row_num}: {str(e)}')
            else:
                # Process Excel file
                workbook = openpyxl.load_workbook(temp_path)
                sheet = workbook.active
                
                # Get headers from first row
                headers = [cell.value for cell in sheet[1]]
                
                for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    if not any(row):  # Skip empty rows
                        continue
                    
                    try:
                        row_dict = dict(zip(headers, row))
                        imported_count += process_question_row(row_dict, row_num)
                    except Exception as e:
                        errors.append(f'Row {row_num}: {str(e)}')
            
            db.session.commit()
            
        finally:
            # Clean up temporary file
            if os.path.exists(temp_path):
                os.remove(temp_path)
        
        response_data = {
            'success': True,
            'imported_count': imported_count,
            'message': f'Successfully imported {imported_count} questions'
        }
        
        if errors:
            response_data['errors'] = errors
            response_data['message'] += f' with {len(errors)} errors'
        
        return jsonify(response_data)
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Import failed: {str(e)}'}), 500

def process_question_row(row, row_num):
    """Process a single row of question data"""
    # Map common column names
    column_mapping = {
        'category': ['category', 'category_name', 'Category', 'Category Name'],
        'question': ['question', 'question_text', 'Question', 'Question Text'],
        'option_a': ['option_a', 'option a', 'Option A', 'A'],
        'option_b': ['option_b', 'option b', 'Option B', 'B'],
        'option_c': ['option_c', 'option c', 'Option C', 'C'],
        'option_d': ['option_d', 'option d', 'Option D', 'D'],
        'correct_answer': ['correct_answer', 'correct', 'answer', 'Correct Answer', 'Answer']
    }
    
    # Extract data using flexible column mapping
    data = {}
    for field, possible_columns in column_mapping.items():
        for col in possible_columns:
            if col in row and row[col]:
                data[field] = str(row[col]).strip()
                break
    
    # Validate required fields
    required_fields = ['category', 'question', 'option_a', 'option_b', 'option_c', 'option_d', 'correct_answer']
    for field in required_fields:
        if field not in data or not data[field]:
            raise ValueError(f'Missing required field: {field}')
    
    # Validate correct answer
    correct_answer = data['correct_answer'].upper()
    if correct_answer not in ['A', 'B', 'C', 'D']:
        raise ValueError('Correct answer must be A, B, C, or D')
    
    # Find or create category
    category = Category.query.filter_by(name=data['category']).first()
    if not category:
        category = Category(name=data['category'])
        db.session.add(category)
        db.session.flush()  # Get the ID without committing
    
    # Create question
    question = Question(
        category_id=category.id,
        question_text=data['question'],
        option_a=data['option_a'],
        option_b=data['option_b'],
        option_c=data['option_c'],
        option_d=data['option_d'],
        correct_answer=correct_answer
    )
    
    db.session.add(question)
    return 1  # Successfully processed one question

@app.route('/admin/api/questions/bulk-import-txt', methods=['POST'])
@login_required
def bulk_import_questions_txt():
    """Bulk import questions from text file with industry-based format"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'No file uploaded'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': 'No file selected'}), 400
        
        if not file.filename.lower().endswith('.txt'):
            return jsonify({'success': False, 'message': 'Only .txt files are supported'}), 400
        
        # Save uploaded file to question_files directory
        filename = secure_filename(file.filename)
        question_files_path = os.path.join('question_files', filename)
        os.makedirs('question_files', exist_ok=True)
        file.save(question_files_path)
        
        # Read file content from saved file
        with open(question_files_path, 'r', encoding='utf-8') as f:
            content = f.read()
        lines = [line.strip() for line in content.split('\n') if line.strip()]
        
        imported_count = 0
        errors = []
        current_industry = None
        question_count_for_industry = 0
        
        i = 0
        while i < len(lines):
            line = lines[i]
            
            # Check if this line is an industry name (starts with INDUSTRY:)
            if line.startswith('INDUSTRY:'):
                # Extract industry name after "INDUSTRY: "
                current_industry = line.replace('INDUSTRY:', '').strip()
                # Remove any numbering like "17. Construction" -> "Construction"
                if '. ' in current_industry:
                    current_industry = current_industry.split('. ', 1)[1]
                question_count_for_industry = 0
                i += 1
                continue
            
            # Process question block (numbered questions like "1. What does...")
            if current_industry and line and line[0].isdigit() and '. ' in line:
                try:
                    # Extract question text (remove number and dot)
                    question_text = line.split('. ', 1)[1].strip()
                    
                    # Get options (next 4 lines)
                    if i + 4 >= len(lines):
                        errors.append(f'Incomplete question block for industry {current_industry}')
                        break
                    
                    option_a = lines[i + 1].replace('A.', '').strip()
                    option_b = lines[i + 2].replace('B.', '').strip()
                    option_c = lines[i + 3].replace('C.', '').strip()
                    option_d = lines[i + 4].replace('D.', '').strip()
                    
                    # Get correct answer (next line)
                    if i + 5 >= len(lines) or not lines[i + 5].startswith('Correct Answer:'):
                        errors.append(f'Missing answer for question: {question_text[:50]}...')
                        i += 5
                        continue
                    
                    correct_answer = lines[i + 5].replace('Correct Answer:', '').strip().upper()
                    
                    # Validate correct answer
                    if correct_answer not in ['A', 'B', 'C', 'D']:
                        errors.append(f'Invalid answer "{correct_answer}" for question: {question_text[:50]}...')
                        i += 6
                        continue
                    
                    # Find or create category for this industry
                    category = Category.query.filter_by(name=current_industry).first()
                    if not category:
                        category = Category(name=current_industry)
                        db.session.add(category)
                        db.session.flush()
                    
                    # Create question
                    question = Question(
                        category_id=category.id,
                        question_text=question_text,
                        option_a=option_a,
                        option_b=option_b,
                        option_c=option_c,
                        option_d=option_d,
                        correct_answer=correct_answer
                    )
                    
                    db.session.add(question)
                    imported_count += 1
                    question_count_for_industry += 1
                    
                    # Move to next question block
                    i += 6
                    
                except Exception as e:
                    errors.append(f'Error processing question in {current_industry}: {str(e)}')
                    i += 6
            else:
                i += 1
        
        db.session.commit()
        
        response_data = {
            'success': True,
            'imported_count': imported_count,
            'message': f'Successfully imported {imported_count} questions from text file'
        }
        
        if errors:
            response_data['errors'] = errors
            response_data['message'] += f' with {len(errors)} errors'
        
        return jsonify(response_data)
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Import failed: {str(e)}'}), 500

@app.route('/admin/api/add-user', methods=['POST'])
@login_required
def add_user():
    data = request.get_json()
    
    user = PreRegisteredUser(
        name=data['name'],
        company_name=data['company_name'],
        industry=data['industry']
    )
    
    db.session.add(user)
    db.session.commit()
    
    return jsonify({'success': True})

@app.route('/admin/api/add-category', methods=['POST'])
@login_required
def add_category():
    data = request.get_json()
    
    category = Category(name=data['name'])
    db.session.add(category)
    db.session.commit()
    
    return jsonify({'success': True})

@app.route('/admin/api/categories', methods=['GET'])
@login_required
def get_categories():
    categories = Category.query.all()
    return jsonify({
        'success': True,
        'categories': [{
            'id': c.id, 
            'name': c.name, 
            'created_at': c.created_at.isoformat(),
            'question_count': len(c.questions)
        } for c in categories]
    })

@app.route('/admin/api/categories/<int:category_id>', methods=['PUT'])
@login_required
def update_category(category_id):
    data = request.get_json()
    category = Category.query.get_or_404(category_id)
    category.name = data['name']
    db.session.commit()
    return jsonify({'success': True})

@app.route('/admin/api/categories/<int:category_id>', methods=['DELETE'])
@login_required
def delete_category(category_id):
    category = Category.query.get_or_404(category_id)
    # Check if category has questions
    if category.questions:
        return jsonify({'success': False, 'message': 'Cannot delete category with existing questions'}), 400
    db.session.delete(category)
    db.session.commit()
    return jsonify({'success': True})

@app.route('/admin/api/bulk-users', methods=['POST'])
@login_required
def bulk_add_users():
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'message': 'No file uploaded'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'message': 'No file selected'}), 400
        
        if not file.filename.lower().endswith(('.csv', '.xlsx', '.xls')):
            return jsonify({'success': False, 'message': 'Invalid file format. Please upload CSV or Excel file'}), 400
        
        # Save uploaded file temporarily
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        users_added = 0
        errors = []
        
        try:
            if filename.lower().endswith('.csv'):
                import csv
                with open(filepath, 'r', encoding='utf-8') as csvfile:
                    reader = csv.DictReader(csvfile)
                    for row_num, row in enumerate(reader, start=2):
                        try:
                            # Check required fields
                            if not all(key in row and row[key].strip() for key in ['name', 'company_name', 'industry']):
                                errors.append(f'Row {row_num}: Missing required fields (name, company_name, industry)')
                                continue
                            
                            user = PreRegisteredUser(
                                name=row['name'].strip(),
                                company_name=row['company_name'].strip(),
                                industry=row['industry'].strip()
                            )
                            db.session.add(user)
                            users_added += 1
                        except Exception as e:
                            errors.append(f'Row {row_num}: {str(e)}')
            
            else:  # Excel file
                workbook = openpyxl.load_workbook(filepath)
                sheet = workbook.active
                
                # Get header row
                headers = [cell.value for cell in sheet[1]]
                required_cols = ['name', 'company_name', 'industry']
                
                # Check if required columns exist
                missing_cols = [col for col in required_cols if col not in headers]
                if missing_cols:
                    return jsonify({
                        'success': False, 
                        'message': f'Missing required columns: {", ".join(missing_cols)}'
                    }), 400
                
                for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        if not row or all(cell is None for cell in row):
                            continue
                        
                        row_data = dict(zip(headers, row))
                        
                        # Check required fields
                        if not all(row_data.get(key) and str(row_data[key]).strip() for key in required_cols):
                            errors.append(f'Row {row_num}: Missing required fields')
                            continue
                        
                        user = PreRegisteredUser(
                            name=str(row_data['name']).strip(),
                            company_name=str(row_data['company_name']).strip(),
                            industry=str(row_data['industry']).strip()
                        )
                        db.session.add(user)
                        users_added += 1
                    except Exception as e:
                        errors.append(f'Row {row_num}: {str(e)}')
            
            db.session.commit()
            
        finally:
            # Clean up temporary file
            if os.path.exists(filepath):
                os.remove(filepath)
        
        return jsonify({
            'success': True,
            'message': f'Successfully added {users_added} users',
            'users_added': users_added,
            'errors': errors
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error processing file: {str(e)}'}), 500

@app.route('/admin/api/bulk-users-txt', methods=['POST'])
@login_required
def bulk_add_users_txt():
    try:
        # Check if all three files are uploaded
        required_files = ['names_file', 'companies_file', 'industries_file']
        for file_key in required_files:
            if file_key not in request.files:
                return jsonify({'success': False, 'message': f'Missing {file_key.replace("_", " ")}'}), 400
            
            file = request.files[file_key]
            if file.filename == '':
                return jsonify({'success': False, 'message': f'No {file_key.replace("_", " ")} selected'}), 400
            
            if not file.filename.lower().endswith('.txt'):
                return jsonify({'success': False, 'message': f'{file_key.replace("_", " ")} must be a .txt file'}), 400
        
        # Create bulk_uploads directory if it doesn't exist
        upload_dir = os.path.join(os.getcwd(), 'bulk_uploads')
        if not os.path.exists(upload_dir):
            os.makedirs(upload_dir)
        
        # Save and process the three files
        names_file = request.files['names_file']
        companies_file = request.files['companies_file']
        industries_file = request.files['industries_file']
        
        # Generate timestamp for unique filenames
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # Save files to bulk_uploads folder
        names_filename = f'names_{timestamp}.txt'
        companies_filename = f'companies_{timestamp}.txt'
        industries_filename = f'industries_{timestamp}.txt'
        
        names_path = os.path.join(upload_dir, names_filename)
        companies_path = os.path.join(upload_dir, companies_filename)
        industries_path = os.path.join(upload_dir, industries_filename)
        
        names_file.save(names_path)
        companies_file.save(companies_path)
        industries_file.save(industries_path)
        
        # Read file contents from saved files
        with open(names_path, 'r', encoding='utf-8') as f:
            names_content = f.read().strip()
        with open(companies_path, 'r', encoding='utf-8') as f:
            companies_content = f.read().strip()
        with open(industries_path, 'r', encoding='utf-8') as f:
            industries_content = f.read().strip()
        
        # Split into lines and clean up
        names = [name.strip() for name in names_content.split('\n') if name.strip()]
        companies = [company.strip() for company in companies_content.split('\n') if company.strip()]
        industries = [industry.strip() for industry in industries_content.split('\n') if industry.strip()]
        
        # Check if all files have the same number of entries
        if len(names) != len(companies) or len(names) != len(industries):
            return jsonify({
                'success': False, 
                'message': f'File length mismatch: Names({len(names)}), Companies({len(companies)}), Industries({len(industries)}). All files must have the same number of entries.'
            }), 400
        
        if len(names) == 0:
            return jsonify({'success': False, 'message': 'All files are empty'}), 400
        
        users_added = 0
        errors = []
        
        # Process each user serially
        for i in range(len(names)):
            try:
                name = names[i]
                company = companies[i]
                industry = industries[i]
                
                # Validate that none are empty
                if not name or not company or not industry:
                    errors.append(f'Line {i+1}: Empty field(s) - Name: "{name}", Company: "{company}", Industry: "{industry}"')
                    continue
                
                # Check if user already exists
                existing_user = PreRegisteredUser.query.filter_by(
                    name=name, 
                    company_name=company, 
                    industry=industry
                ).first()
                
                if existing_user:
                    errors.append(f'Line {i+1}: User "{name}" from "{company}" in "{industry}" already exists')
                    continue
                
                # Create new user
                user = PreRegisteredUser(
                    name=name,
                    company_name=company,
                    industry=industry
                )
                db.session.add(user)
                users_added += 1
                
            except Exception as e:
                errors.append(f'Line {i+1}: {str(e)}')
        
        # Commit all changes
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Successfully added {users_added} users',
            'users_added': users_added,
            'total_processed': len(names),
            'errors': errors,
            'saved_files': {
                'names_file': names_path,
                'companies_file': companies_path,
                'industries_file': industries_path
            }
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Error processing files: {str(e)}'}), 500

@app.route('/admin/api/dashboard-stats', methods=['GET'])
@login_required
def get_dashboard_stats():
    """Get comprehensive dashboard statistics"""
    try:
        from datetime import datetime, timedelta
        from sqlalchemy import func, text
        
        today = datetime.now().date()
        
        # Basic counts
        total_users = User.query.count()
        total_questions = Question.query.count()
        total_games = GameSession.query.count()
        total_categories = Category.query.count()
        
        # Today's counts
        users_today = User.query.filter(func.date(User.created_at) == today).count()
        games_today = GameSession.query.filter(func.date(GameSession.created_at) == today).count()
        
        # Average success rate calculation (percentage of correct answers)
        total_answers = GameSession.query.count()
        correct_answers = GameSession.query.filter_by(is_correct=True).count()
        avg_score = round((correct_answers / total_answers * 100), 1) if total_answers > 0 else 0
        
        # Question statistics by category
        category_stats = db.session.query(
            Category.name,
            func.count(Question.id).label('question_count')
        ).outerjoin(Question).group_by(Category.id, Category.name).all()
        
        # Most difficult questions (lowest success rate)
        difficult_questions = db.session.query(
            Question.question,
            Category.name.label('category_name'),
            func.count(GameSession.id).label('attempts'),
            func.avg(func.case([(GameSession.is_correct == True, 1)], else_=0)).label('success_rate')
        ).join(Category).outerjoin(GameSession).group_by(Question.id).having(
            func.count(GameSession.id) > 0
        ).order_by(text('success_rate ASC')).limit(5).all()
        
        # Recent activity (last 7 days)
        week_ago = today - timedelta(days=7)
        recent_games = GameSession.query.filter(
            GameSession.created_at >= week_ago
        ).count()
        
        stats = {
            'total_users': total_users,
            'total_questions': total_questions,
            'total_games': total_games,
            'total_categories': total_categories,
            'users_today': users_today,
            'games_today': games_today,
            'avg_score': avg_score,
            'recent_games': recent_games,
            'category_stats': [{
                'name': stat.name,
                'question_count': stat.question_count
            } for stat in category_stats],
            'difficult_questions': [{
                'question': q.question[:100] + '...' if len(q.question) > 100 else q.question,
                'category': q.category_name,
                'attempts': q.attempts,
                'success_rate': round(float(q.success_rate) * 100, 1) if q.success_rate else 0
            } for q in difficult_questions]
        }
        
        return jsonify({
            'success': True,
            'stats': stats
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/admin/api/question-analytics', methods=['GET'])
@login_required
def get_question_analytics():
    """Get detailed question analytics"""
    try:
        from sqlalchemy import func, text
        
        # Question performance analytics
        question_performance = db.session.query(
            Question.id,
            Question.question,
            Category.name.label('category_name'),
            func.count(GameSession.id).label('total_attempts'),
            func.sum(func.case([(GameSession.is_correct == True, 1)], else_=0)).label('correct_attempts'),
            func.avg(func.case([(GameSession.is_correct == True, 100)], else_=0)).label('avg_score_when_answered')
        ).join(Category).outerjoin(GameSession).group_by(
            Question.id, Question.question, Category.name
        ).all()
        
        # Category performance
        category_performance = db.session.query(
            Category.name,
            func.count(Question.id).label('total_questions'),
            func.count(GameSession.id).label('total_attempts'),
            func.avg(GameSession.score).label('avg_score')
        ).outerjoin(Question).outerjoin(GameSession).group_by(
            Category.id, Category.name
        ).all()
        
        # Format question performance data
        questions_data = []
        for q in question_performance:
            success_rate = (q.correct_attempts / q.total_attempts * 100) if q.total_attempts > 0 else 0
            questions_data.append({
                'id': q.id,
                'question': q.question[:100] + '...' if len(q.question) > 100 else q.question,
                'category': q.category_name,
                'total_attempts': q.total_attempts or 0,
                'correct_attempts': q.correct_attempts or 0,
                'success_rate': round(success_rate, 1),
                'avg_score': round(q.avg_score_when_answered, 1) if q.avg_score_when_answered else 0,
                'difficulty_level': 'Easy' if success_rate > 70 else 'Medium' if success_rate > 40 else 'Hard'
            })
        
        # Format category performance data
        categories_data = [{
            'name': c.name,
            'total_questions': c.total_questions or 0,
            'total_attempts': c.total_attempts or 0,
            'avg_score': round(c.avg_score, 1) if c.avg_score else 0
        } for c in category_performance]
        
        return jsonify({
            'success': True,
            'question_performance': questions_data,
            'category_performance': categories_data
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/admin/api/reset-sessions', methods=['POST'])
@login_required
def reset_sessions():
    """Reset all game sessions and user journeys while keeping user data, questions, and other information intact"""
    try:
        # Clear session-related data and user journeys
        GameSession.query.delete()
        UserJourney.query.delete()
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'All game sessions and user journeys have been reset successfully'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/admin/api/excel-report', methods=['GET'])
@login_required
def download_excel_report():
    """Generate and download complete Excel report with user data"""
    try:
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Game Report"
        
        # Define headers
        headers = ['User Name', 'Time of Playing', 'Company Name', 'Industry', 'Email ID', 'Phone', 'Job Title', 'Department', 'Selfie Filename']
        
        # Add headers with styling
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Get all completed user journeys in chronological order (most recent first)
        journeys = UserJourney.query.filter_by(is_completed=True).order_by(UserJourney.journey_start.desc()).all()
        
        # Set row height for images
        ws.row_dimensions[1].height = 20  # Header row
        
        row = 2
        for journey in journeys:
            # Add user data directly from journey (all data is stored in UserJourney table)
            ws.cell(row=row, column=1, value=journey.name or 'N/A')
            ws.cell(row=row, column=2, value=journey.journey_start.strftime('%Y-%m-%d %H:%M:%S') if journey.journey_start else 'N/A')
            ws.cell(row=row, column=3, value=journey.company_name or 'N/A')
            ws.cell(row=row, column=4, value=journey.industry or 'N/A')
            ws.cell(row=row, column=5, value=journey.email or 'N/A')
            ws.cell(row=row, column=6, value=journey.phone or 'N/A')
            ws.cell(row=row, column=7, value=journey.job_title or 'N/A')
            ws.cell(row=row, column=8, value=journey.department or 'N/A')
            
            # Handle selfie image from journey - use filename only to avoid Excel corruption
            if journey.selfie_filename:
                try:
                    # Just store the filename instead of embedding the image
                    # This prevents Excel file corruption issues
                    selfie_path = os.path.join('static', 'uploads', 'selfies', journey.selfie_filename)
                    
                    if os.path.exists(selfie_path):
                        ws.cell(row=row, column=9, value=journey.selfie_filename)
                    else:
                        ws.cell(row=row, column=9, value='Image File Missing')
                except Exception as img_error:
                    ws.cell(row=row, column=9, value='Error accessing image')
            else:
                ws.cell(row=row, column=9, value='No Selfie')
            
            row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
                
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename with current date
        filename = f"complete-game-report-{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500